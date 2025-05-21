from openpyxl import Workbook, load_workbook
#ask for message input function
username = input("Jūsu vārds ")
print("Sveiki", username,"!")

# 'path' var pierakstīt maršrutu excel failam
path = r"C:\Users\excel\OneDrive\Documents\projekts python\pieraksti.xlsx"

tasks = {
    "do homework": "23:00",
    "watch movie": "23:30"
}


def load(filename=path):
    tasks = {}
    wb=load_workbook(filename) 
    ws=wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        task, time = row
        if task and time:
                tasks[task] = time
    return tasks

def save_file(filename=path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws['A1'] = "Uzdevumi"
    ws['B1'] = "Laiks"
    for row, (task, time) in enumerate(tasks.items(), start=2): 
        ws[f'A{row}'] = task
        ws[f'B{row}'] = time
    wb.save(filename)

def add_task():
    task = input("ievadiet jaunu uzdevumu ")
    #tasks.append(name)
    print("Uzdevums " + task.lower() + " ir pievienots")
    save_file()

def add_time():
    time = input("ievadiet laiku kad jāpabeidz: st:min ")
    #tasks.append({"time": time}) works for lists not dictionaries
    hours, minutes = convert(time)
    tasks[tasks] = time
    print(hours)
    print(minutes, " laiks pievienots")
    save_file()
    
def add_both():
    task = input("Ievadiet uzdevumu ")
    #tasks.append({"name": name, "time": time}) works for lists
    time = input("Ievadiet laiku uzdevumam formātā 'stundas' : 'minūtes'")
    hours, minutes = convert(time)
    tasks[task] = time
    print(" pievienots!")
    save_file()
    
    
def remove_task():
    name = input("ko nodzēst? ")
    if name in tasks:
        #tasks.remove(name)
        del tasks[name]
        print("izdzēsts")
    else:
        print("sorry, nav atrasts :)")
    save_file()

def print_time():
    for i in tasks:
        print(tasks[i])
    
def print_tasks():
    for task in tasks:
        print(task)
        
def print_both():
    for task in tasks:
        print(task)
        print(tasks[task])
def main():
    tasks = load()
    print("Kas ir jāizdara? Add / Del / See / Out")
#lowercase str.lower()
### save data in dict?
    while True:
        do = input().lower().strip()
        if do == "add":
            add_both()
        elif do == "del":
            remove_task()
        elif do == "see":
            print_both()
        elif do == "out":
            print("Paldies!")
            break
        else:
            print("ko?")

###txt.replace()   changes task to "task1 " for example

def convert(time):
    hours, minutes = map(int, time.split(":"))
    if (hours >= 0 and hours) <= 23 and (minutes >= 0 and minutes <= 59):
        return hours, minutes
    
main()
